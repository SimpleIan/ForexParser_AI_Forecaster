import requests
from bs4 import BeautifulSoup
import openpyxl
import os

# Parser

def priceTrackerYahoo():
    url = 'https://finance.yahoo.com/quote/EURUSD=X?p=EURUSD=X&.tsrc=fin-srch'
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'lxml')

    YahooPrice = soup.find_all('div', {'class': 'My(6px) Pos(r) smartphone_Mt(6px)'})[0].find('span').text
    return YahooPrice

# def priceTracerstooq():
#     url = 'https://stooq.com/q/d/?s=eurusd&c=0&d1=20200601&d2=20200630'
#     response = requests.get(url)
#     soup = BeautifulSoup(response.text, 'lxml')
#
#     stooprice = soup.find({'span'}, {'id': 'aq_eurusd_c5'}).text
#     return stooprice
#
# def priceTracerFinam():
#     url = 'https://www.finam.ru/quote/forex/eur-usd/'
#     response = requests.get(url)
#     soup = BeautifulSoup(response.text, 'lxml')
#
#     finamprice = soup.find({'span'}, {'class': 'PriceInformation__price--26G'}).text
#     return finamprice

# excelWriter
# create a new Excell file

while True:

    print('YahooPrice:'+priceTrackerYahoo())
    filename = 'Stocks.xlsx'
    if os.path.exists(filename):
        xlsfile = openpyxl.load_workbook(filename)
    else:
        xlsfile = openpyxl.Workbook()

    xlsfile.sheetnames
    sheet = xlsfile['Sheet']

    #Yahoo add data
    data = [priceTrackerYahoo()]
    max_rows = sheet.max_row
    for row, (data) in enumerate(data, start=1):
        sheet['A{}'.format(row + max_rows)].value = priceTrackerYahoo()
   # for row, (data) in enumerate(data, start=0):
   #     sheet['B{}'.format(row + max_rows)].value = priceTracerstooq()
   # for row, (data) in enumerate(data, start=0):
   #     sheet['C{}'.format(row + max_rows)].value = priceTracerFinam()
    #save
    xlsfile.save('Stocks.xlsx')


